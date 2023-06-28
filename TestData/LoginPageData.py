import openpyxl


class HomePageData:
    test_HomePage_data = [{"GID": "000006", "Password": "Password@123"},
                          {"GID": "000607", "Password": "Password@123"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(r"C:\\Users\\naveen.kstb\\Desktop\\AssessLogin.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # To get Rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(1, sheet.max_column + 1):  # To get Columns
                    # print(sheet.cell(row=i, column=j).value)

                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]